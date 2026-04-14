import openpyxl
import sys

file_path = r'C:\Users\haptn\Downloads\Ha.Odoo\System-Analysis Template.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
except PermissionError:
    print("PERMISSION ERROR: The file is currently open in Excel. Please close it before running.")
    sys.exit(1)
except Exception as e:
    print(f"Error loading workbook: {e}")
    sys.exit(1)

def ensure_unmerged(ws, row, col):
    for merged_range in list(ws.merged_cells.ranges):
        if row >= merged_range.min_row and row <= merged_range.max_row and col >= merged_range.min_col and col <= merged_range.max_col:
            ws.unmerge_cells(str(merged_range))

def set_cell(ws, row, col, val):
    ensure_unmerged(ws, row, col)
    ws.cell(row=row, column=col, value=val)

def get_last_row(sheet, col_idx):
    for r in range(sheet.max_row, 0, -1):
        # We can't safely read .value if it's a MergedCell, but openpyxl handles reading it as None most times unless we access it directly
        cell = sheet.cell(row=r, column=col_idx)
        if type(cell).__name__ != 'MergedCell' and cell.value is not None:
            return r
    return 1 # Fallback

# 1. Actor List
ws_actor = wb['Actor List']
last_row = get_last_row(ws_actor, 3) # C (Actor Name)
start_row = max(last_row + 1, 3)
actors = [
    [1, 'Salesperson', 'Chịu trách nhiệm chính cho việc tìm kiếm, chăm sóc khách hàng ở phân hệ CRM và tạo báo giá, chốt đơn hàng ở phân hệ Sales.'],
    [2, 'Sales Manager', 'Quản lý đội ngũ bán hàng, theo dõi toàn bộ pipeline và phê duyệt các ngoại lệ hoặc chính sách giảm giá (nếu có).']
]
for p in actors:
    set_cell(ws_actor, start_row, 2, p[0])
    set_cell(ws_actor, start_row, 3, p[1])
    set_cell(ws_actor, start_row, 4, p[2])
    start_row += 1

# 2. Function List
ws_func = wb['Function List']
last_row = get_last_row(ws_func, 4) # D (Function)
start_row = max(last_row + 1, 3)
funcs = [
    [1, 'Opportunity', 'Create Opportunity', 'Khởi tạo một cơ hội bán hàng mới trong Pipeline.', 'Basic'],
    [2, 'Opportunity', 'Move Stage', 'Chuyển đổi trạng thái của cơ hội (New -> Qualified -> Proposition -> Won/Lost).', 'Basic'],
    [3, 'Opportunity', 'Mark as Won / Lost', 'Chốt kết quả cơ hội bán hàng (Thắng/Thua).', 'Process'],
    [4, 'Quotation', 'Create Quotation', 'Tạo một báo giá mới từ cơ hội bán hàng hiện tại, kế thừa thông tin khách hàng.', 'Basic'],
    [5, 'Quotation', 'Add Order Line', 'Thêm sản phẩm, số lượng, và giá vào báo giá.', 'Basic'],
    [6, 'Quotation', 'Send by Email', 'Gửi báo giá cho khách hàng qua email.', 'Operation'],
    [7, 'Sales Order', 'Confirm Order', 'Xác nhận khách hàng đồng ý báo giá và chuyển báo giá thành Đơn bán hàng.', 'Process']
]
for p in funcs:
    set_cell(ws_func, start_row, 2, p[0])
    set_cell(ws_func, start_row, 3, p[1])
    set_cell(ws_func, start_row, 4, p[2])
    set_cell(ws_func, start_row, 5, p[3])
    set_cell(ws_func, start_row, 6, p[4])
    start_row += 1

# 3. Workflow
ws_wf = wb['Workflow']
last_row = get_last_row(ws_wf, 4) # D (Function)
start_row = max(last_row + 1, 3)
wfs = [
    ['(None)', 'Salesperson', 'Create Opportunity', 'Opportunity', 'New', 'Bắt đầu tạo mới khi có leads/khách hàng tiềm năng.'],
    ['New', 'Salesperson', 'Move Stage', 'Opportunity', 'Qualified', 'Sau khi đã thẩm định thông tin khách hàng.'],
    ['Qualified', 'Salesperson', 'Move Stage', 'Opportunity', 'Proposition', 'Đang trong giai đoạn chuẩn bị hoặc trình bày giải pháp.'],
    ['Proposition', 'Salesperson', 'Create Quotation', 'Quotation', 'Draft', 'Xác định nhu cầu mua cụ thể và lập báo giá.'],
    ['Draft', 'Salesperson', 'Send by Email', 'Quotation', 'Quotation Sent', 'Gửi cho khách hàng xem xét.'],
    ['Quotation Sent', 'Salesperson', 'Confirm Order', 'Sales Order', 'Sale Order', 'Khách hàng đồng ý, hệ thống chốt giao dịch.'],
    ['Proposition', 'Salesperson', 'Mark as Won', 'Opportunity', 'Won', 'Có thể đánh dấu tự động khi Confirm Order.']
]
for idx, p in enumerate(wfs):
    set_cell(ws_wf, start_row, 1, (idx+1))
    set_cell(ws_wf, start_row, 2, p[0])
    set_cell(ws_wf, start_row, 3, p[1])
    set_cell(ws_wf, start_row, 4, p[2])
    set_cell(ws_wf, start_row, 5, p[3])
    set_cell(ws_wf, start_row, 6, p[4])
    set_cell(ws_wf, start_row, 7, p[5])
    start_row += 1

# 4. Data Object Details
ws_do = wb['Data Object Details']
last_row = get_last_row(ws_do, 3) # C (Data Object)
start_row = max(last_row + 1, 3)
dos = [
    ['Salesperson', 'Opportunity', 'Cơ hội bán hàng', 'Opportunity Name', 'Nhu cầu mua 5 phần mềm Odoo', 'Text box', 'Yes', 'Yes'],
    ['Salesperson', 'Opportunity', 'Cơ hội bán hàng', 'Customer', 'Azure Interior', 'Dropdown/Lookup', 'Yes', 'Yes'],
    ['Salesperson', 'Opportunity', 'Cơ hội bán hàng', 'Expected Revenue', '50,000,000', 'Number / Currency', 'Yes', 'Yes'],
    ['Salesperson', 'Sales Order', 'Đơn bán hàng (Báo giá)', 'Expiration Date', '15/04/2026', 'Date Picker', 'Yes', 'Yes'],
    ['Salesperson', 'Sales Order', 'Đơn bán hàng (Báo giá)', 'Order Lines', '[Danh sách SP, SL, Đơn giá]', 'Grid / Table', 'Yes', 'Yes'],
    ['Salesperson', 'Sales Order', 'Đơn bán hàng (Báo giá)', 'Total Amount', '55,000,000', 'Label (Auto-calc)', 'Yes', 'Yes']
]
for idx, p in enumerate(dos):
    set_cell(ws_do, start_row, 1, (idx+1))
    set_cell(ws_do, start_row, 2, p[0])
    set_cell(ws_do, start_row, 3, p[1])
    set_cell(ws_do, start_row, 4, p[2])
    set_cell(ws_do, start_row, 5, p[3])
    set_cell(ws_do, start_row, 6, p[4])
    set_cell(ws_do, start_row, 7, p[5])
    set_cell(ws_do, start_row, 8, p[6])
    set_cell(ws_do, start_row, 9, p[7])
    start_row += 1


# 5. Screen List
ws_sc = wb['Screen List']
last_row = get_last_row(ws_sc, 4) # D (Screen)
start_row = max(last_row + 1, 3)
scs = [
    ['CRM', 'CRM Pipeline Kanban', 'Hiển thị quá trình phân loại các cơ hội bán hàng theo dạng cột (Kanban tab).', 'To-do'],
    ['CRM', 'Opportunity Form', 'Giao diện chi tiết thông tin một cơ hội bán hàng cụ thể.', 'To-do'],
    ['Sales', 'Quotations List', 'Danh sách liệt kê toàn bộ các báo giá trong hệ thống.', 'To-do'],
    ['Sales', 'Sales Order Form', 'Giao diện chi tiết của một báo giá / đơn hàng, bao gồm các order lines.', 'To-do']
]
for idx, p in enumerate(scs):
    set_cell(ws_sc, start_row, 2, (idx+1))
    set_cell(ws_sc, start_row, 3, p[0])
    set_cell(ws_sc, start_row, 4, p[1])
    set_cell(ws_sc, start_row, 5, p[2])
    set_cell(ws_sc, start_row, 6, p[3])
    start_row += 1

new_file_path = r'C:\Users\haptn\Downloads\Ha.Odoo\Ha-Odoo- System-Analysis.xlsx'
try:
    wb.save(new_file_path)
    print("SUCCESS: Saved to new file.")
except Exception as e:
    print(f"Error saving workbook to new file: {e}")

